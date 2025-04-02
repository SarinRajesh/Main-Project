from django.shortcuts import render, redirect, HttpResponse, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.core.mail import send_mail
from django.urls import reverse
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.auth.hashers import make_password
from .models import UserType, Consultation, Users, Design, Amount, Product, Cart, Review, Order, Payment_Type, ConsultationDate,VirtualRoom, RoomModel, VirtualRoomModel
from .decorators import nocache
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from allauth.socialaccount.models import SocialAccount
from django.views.decorators.csrf import csrf_exempt, csrf_protect
import random
from django.utils.dateparse import parse_datetime, parse_date
from decimal import Decimal, InvalidOperation
from django.views.decorators.http import require_POST, require_http_methods
from django.utils import timezone
from django.middleware.csrf import get_token
from django.db.models import Sum, Count
from django.db import transaction
from django.core.exceptions import ValidationError
from allauth.account.adapter import DefaultAccountAdapter
from allauth.socialaccount.adapter import DefaultSocialAccountAdapter
import json
import re
from datetime import datetime, time
from django.views.decorators.cache import never_cache
from django.utils.formats import date_format
from datetime import timedelta
import google.generativeai as genai
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
import json
from django.http import JsonResponse
from django.core.serializers import serialize
from django.forms.models import model_to_dict
from django.core.files.storage import default_storage
from django.http import HttpResponseForbidden
from xhtml2pdf import pisa




class CustomTokenGenerator(PasswordResetTokenGenerator):
    def _make_hash_value(self, user, timestamp):
        return str(user.pk) + str(timestamp)

token_generator = CustomTokenGenerator()


@nocache
def index(request):
    context = {}
    if request.user.is_authenticated:
        user = request.user
        user_type = user.user_type_id.user_type if user.user_type_id else None
        context = {
            'user': user, 
            'user_type': user_type,
            'is_authenticated': True
        }
    else:
        context = {
            'is_authenticated': False
        }
    return render(request, 'index.html', context)



def signin(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.status == 'active':  # Check if user is active
                login(request, user)  # Log in the user
                return custom_login_redirect(request)  # Redirect to the appropriate page
            else:
                # Store message in session under a custom key
                request.session['custom_error_message'] = "User account is inactive."
                return redirect('signin')  # Redirect to sign-in page
        else:
            # Store message in session under a custom key
            request.session['custom_error_message'] = "Invalid username or password."
            return redirect('signin')  # Redirect to sign-in page
    
    return render(request, 'signin.html')


def custom_login_redirect(request):
    user = request.user
    if not user.is_authenticated:
        return redirect('signin')

    if user.status != 'active':  # Check if user is active
        logout(request)  # Log out inactive users
        # Store message in session under a custom key
        request.session['custom_error_message'] = "User account is inactive."
        return redirect('signin')  # Redirect to sign-in page

    # Check user type and redirect accordingly
    if user.user_type_id:
        if user.user_type_id.user_type == 'Admin':
            return redirect('admin_index')
        elif user.user_type_id.user_type == 'Delivery_boy':
            return redirect('deliveryboy_index')  # Changed from 'delivery_boy/index' to 'deliveryboy_index'
    
    # Default redirect for other user types (e.g., Customer, Designer)
    return redirect('index')



def generate_otp():
    return random.randint(100000, 999999)

class CustomAccountAdapter(DefaultAccountAdapter):
    def save_user(self, request, user, form, commit=True):
        user = super().save_user(request, user, form, commit=False)
        user.user_type_id = UserType.objects.get(user_type='Customer')
        if commit:
            user.save()
        return user

# Update the CustomSocialAccountAdapter class
class CustomSocialAccountAdapter(DefaultSocialAccountAdapter):
    def pre_social_login(self, request, sociallogin):
        # Check if the email already exists in the database
        email = sociallogin.account.extra_data['email']
        if email:
            try:
                user = Users.objects.get(email=email)
                # If the user exists, connect the social account to the existing user
                sociallogin.connect(request, user)
            except Users.DoesNotExist:
                pass  # If the user doesn't exist, proceed with the normal flow

    def save_user(self, request, sociallogin, form=None):
        user = super().save_user(request, sociallogin, form)
        user.user_type_id = UserType.objects.get(user_type='Customer')
        user.save()
        return user

# Update the signup function
def signup(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        phone = request.POST.get('phone')
        email = request.POST.get('email')
        username = request.POST.get('username')
        password = request.POST.get('password')

        # Check if a user with this email already exists
        existing_user = Users.objects.filter(email=email).first()
        if existing_user:
            # If the user exists, check if they have a social account
            social_account = SocialAccount.objects.filter(user=existing_user).first()
            if social_account:
                messages.error(request, "An account with this email already exists. Please use Google login.")
                return redirect('signin')

        # Store user data in session
        request.session['user_data'] = {
            'name': name,
            'phone': phone,
            'email': email,
            'username': username,
            'password': password,
        }
        
        otp = generate_otp()
        request.session['otp'] = otp

        # Send OTP to user's email
        send_mail(
            'Your OTP for signing up',
            f'Your OTP is {otp}',
            'your-email@gmail.com',  # Replace with your actual email
            [email],
            fail_silently=False,
        )

        messages.success(request, "Check your email for the OTP.")
        return redirect('verify_otp')
        
    return render(request, 'signup.html')

# Update the verify_otp function
def verify_otp(request):
    if request.method == 'POST':
        otp = request.POST.get('otp')
        session_otp = request.session.get('otp')

        if otp == str(session_otp):
            user_data = request.session.get('user_data')
            if user_data:
                try:
                    # Check if a user with this email already exists
                    existing_user = Users.objects.filter(email=user_data['email']).first()
                    if existing_user:
                        # If the user exists, update their information
                        existing_user.name = user_data['name']
                        existing_user.phone = user_data['phone']
                        existing_user.username = user_data['username']
                        existing_user.password = make_password(user_data['password'])
                        existing_user.save()
                        user = existing_user
                    else:
                        # Create a new user with the provided data
                        user = Users.objects.create(
                            name=user_data['name'],
                            phone=user_data['phone'],
                            email=user_data['email'],
                            username=user_data['username'],
                            password=make_password(user_data['password']),
                            user_type_id=UserType.objects.get(user_type='Customer')
                        )

                    # Log the user in
                    backend = 'django.contrib.auth.backends.ModelBackend'
                    user = authenticate(username=user_data['username'], password=user_data['password'], backend=backend)
                    if user is not None:
                        login(request, user)
                    
                    # Clear user data from session
                    request.session.pop('user_data', None)

                    return redirect('index')
                except Exception as e:
                    messages.error(request, f"Error creating account: {str(e)}")
            else:
                messages.error(request, "User data not found in session. Please sign up again.")
                return redirect('signup')
        else:
            messages.error(request, "Invalid OTP. Please try again.")

    return render(request, 'verify_otp.html')


def check_email(request):
    email = request.GET.get('email', None)
    data = {
        'is_taken': Users.objects.filter(email=email).exists()
    }
    return JsonResponse(data)

def check_username(request):
    username = request.GET.get('username', None)
    data = {
        'is_taken': Users.objects.filter(username=username).exists()
    }
    return JsonResponse(data)

def request_password_reset(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = Users.objects.get(email=email)
            
            # Generate a password reset token and send it to the user's email
            token = token_generator.make_token(user)
            uidb64 = urlsafe_base64_encode(force_bytes(user.pk))
            
            reset_url = reverse('reset_password_confirm', kwargs={'uidb64': uidb64, 'token': token})
            reset_url = request.build_absolute_uri(reset_url)
            
            send_mail(
                'Password Reset Request',
                f'Click the following link to reset your password: {reset_url}',
                'your-email@gmail.com',  # Replace with your actual email
                [email],
                fail_silently=False,
            )
            
            messages.success(request, 'Password reset email has been sent. Check your email to proceed.')
            return redirect('request_password_reset')
        
        except Users.DoesNotExist:
            messages.error(request, 'User does not exist.')
            return redirect('request_password_reset')

    return render(request, 'password_reset/request_password_reset.html')

def reset_password_confirm(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = Users.objects.get(pk=uid)
        
        if token_generator.check_token(user, token):
            if request.method == 'POST':
                new_password = request.POST.get('new_password')
                confirm_password = request.POST.get('confirm_password')
                
                if new_password == confirm_password:
                    user.password = make_password(new_password)  # Hash the password
                    user.save()
                    messages.success(request, 'Password reset successful. You can now sign in with your new password.')
                    return redirect('signin')
                else:
                    messages.error(request, 'Passwords do not match. Please try again.')
            return render(request, 'password_reset/reset_password_confirm.html', {'uidb64': uidb64, 'token': token})
        else:
            messages.error(request, 'Invalid password reset link.')
            return redirect('signin')
    
    except (TypeError, ValueError, OverflowError, Users.DoesNotExist):
        user = None
    
    return redirect('signin')

def logout_view(request):
    logout(request)
    return redirect('signin')



@nocache
@login_required
def profile(request):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None
    return render(request, 'profile.html', {'user': user, 'user_type': user_type})

@login_required
def edit_profile(request):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None

    if request.method == 'POST':
        # Update user profile fields
        user.name = request.POST.get('name')
        user.phone = request.POST.get('phone')
        user.username = request.POST.get('username')
        user.email = request.POST.get('email')
        user.address = request.POST.get('address')
        user.home_town = request.POST.get('home_town')
        user.district = request.POST.get('district')
        user.state = request.POST.get('state')
        user.pincode = request.POST.get('pincode')

        # Handle photo upload if present
        if 'photo' in request.FILES:
            user.photo = request.FILES['photo']

        try:
            user.save()
            messages.success(request, 'Profile updated successfully.')
            return redirect('profile')
        except Exception as e:
            messages.error(request, f'Error updating profile: {str(e)}')
            return redirect('edit_profile')

    context = {
        'user': user,
        'user_type': user_type
    }
    return render(request, 'edit_profile.html', context)



@login_required
@csrf_exempt
def upload_photo(request):
    if request.method == 'POST':
        photo = request.FILES.get('photo')
        if photo:
            user = request.user
            user.photo = photo
            user.save()
            return JsonResponse({'success': True, 'photo_url': user.photo.url})
    return JsonResponse({'success': False})



@login_required
@nocache
def add_product(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        amount_value = request.POST.get('amount')
        image = request.FILES.get('image')
        category = request.POST.get('category')
        stock = request.POST.get('stock')
        color = request.POST.get('color')  # Add this line to get the color value

        amount = Amount(amount=amount_value)
        amount.save()

        product = Product(
            name=name,
            description=description,
            amount=amount,
            category=category,
            image=image,
            stock=stock,
            color=color  # Add this line to save the color
        )
        product.save()

        return redirect('add_product')

    return render(request, 'admin_page/add_product.html')

@login_required
@nocache
def add_portfolio(request):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None

    # Check if user details are complete
    user_details_complete = all([
        user.name,
        user.email,
        user.phone,
        user.address,
        user.home_town,
        user.district,
        user.state,
        user.pincode
    ])

    if request.method == 'POST':
        if not user_details_complete:
            messages.error(request, 'Please complete your profile before adding a portfolio.')
            return redirect('profile')

        name = request.POST.get('name')
        description = request.POST.get('description')
        amount_value = request.POST.get('amount')
        image = request.FILES.get('image')
        category = request.POST.get('category')
        sqft = request.POST.get('sqft', 0)

        amount = Amount(amount=amount_value)
        amount.save()

        portfolio = Design(
            designer_id=request.user,
            name=name,
            description=description,
            amount=amount,
            image=image,
            category=category,
            sqft=sqft
        )
        portfolio.save()

        messages.success(request, 'Portfolio added successfully.')
        return redirect('portfolio')

    context = {
        'user_type': user_type,
        'user_details_complete': user_details_complete
    }
    return render(request, 'add_portfolio.html', context)


from django.http import JsonResponse
from django.template.loader import render_to_string
@nocache
def portfolio(request):
    category = request.GET.get('category', 'all')
    sqft_range = request.GET.get('sqft_range', 'all')
    designer = request.GET.get('designer', 'all')
    search_query = request.GET.get('search', '')

    if request.user.is_authenticated and request.user.user_type_id.user_type == 'Designer':
        # For designers, show only their designs
        portfolios = Design.objects.filter(designer_id=request.user)
    else:
        # For other users, show all designs
        portfolios = Design.objects.all()

    if search_query:
        portfolios = portfolios.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(designer_id__username__icontains=search_query)
        )

    if category != 'all':
        portfolios = portfolios.filter(category__iexact=category)

    if designer != 'all':
        portfolios = portfolios.filter(designer_id__username=designer)

    if sqft_range != 'all':
        sqft_ranges = {
            '0-500': (0, 500),
            '501-1000': (501, 1000),
            '1001-1500': (1001, 1500),
            '1501+': (1501, float('inf'))
        }
        min_sqft, max_sqft = sqft_ranges.get(sqft_range, (None, None))
        if min_sqft is not None:
            portfolios = portfolios.filter(sqft__gte=min_sqft)
        if max_sqft != float('inf'):
            portfolios = portfolios.filter(sqft__lte=max_sqft)

    all_designers = Users.objects.filter(user_type_id__user_type='Designer').values_list('username', flat=True).distinct()
    all_categories = Design.objects.values_list('category', flat=True).distinct()

    def format_category(category):
        return category.replace('_', ' ').title()

    formatted_categories = [{'value': cat, 'display': format_category(cat)} for cat in all_categories]

    context = {
        'portfolios': portfolios,
        'designers': all_designers,
        'categories': formatted_categories,
        'no_results': portfolios.count() == 0,
        'selected_category': category,
        'selected_designer': designer,
        'selected_sqft_range': sqft_range,
        'search_query': search_query,
    }

    if request.user.is_authenticated:
        user = request.user
        user_type = user.user_type_id.user_type if user.user_type_id else None
        context['user_type'] = user_type

    return render(request, 'portfolio.html', context)

@login_required
def delete_portfolio(request, portfolio_id):
    # Get the design item or return 404 if not found
    design = get_object_or_404(Design, id=portfolio_id)
    
    # Check if the current user is the owner of the design item
    if request.user == design.designer_id:
        # Delete the design item
        design.delete()
        messages.success(request, "Portfolio item deleted successfully.")
    else:
        messages.error(request, "You don't have permission to delete this portfolio item.")
    
    # Redirect to the portfolio list page or wherever you want
    return redirect('portfolio')  # Make sure 'portfolio' is the correct name for your portfolio list view

@nocache
def shop(request):
    products = Product.objects.all()
    
    # Get filter parameters
    category = request.GET.get('category')
    price_order = request.GET.get('price_order')
    
    # Filter by category
    if category and category != 'all':
        products = products.filter(category=category)
    
    # Order by price
    if price_order == 'low_to_high':
        products = products.order_by('amount__amount')
    elif price_order == 'high_to_low':
        products = products.order_by('-amount__amount')
    
    # Get unique categories for the dropdown
    categories = Product.objects.values_list('category', flat=True).distinct()
    
    context = {
        'products': products,
        'categories': categories,
        'selected_category': category,
        'selected_price_order': price_order,
    }

    if request.user.is_authenticated:
        user = request.user
        user_type = user.user_type_id.user_type if user.user_type_id else None
        context['user_type'] = user_type
        context['user'] = user

        # Get cart items for the current user
        cart_items = Cart.objects.filter(user_id=user)
        context['cart_items'] = cart_items
        context['cart_total'] = sum(item.amount for item in cart_items)
        context['cart_count'] = cart_items.count()

    return render(request, 'shop.html', context)

from django.shortcuts import render, get_object_or_404


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from .models import Design, Consultation, ConsultationDate
from django.contrib import messages
from django.shortcuts import redirect
from django.utils import timezone

import logging

logger = logging.getLogger(__name__)

@login_required
def portfolio_details(request, portfolio_id):
    portfolio = get_object_or_404(Design.objects.select_related('designer_id'), id=portfolio_id)
    
    # Replace underscores with spaces in the category
    portfolio.category_display = portfolio.category.replace("_", " ")
    
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None
    user_details_complete = all([
        user.address,
        user.home_town,
        user.district,
        user.state,
        user.pincode
    ])

    # Get the latest consultation for this design and user
    latest_consultation = Consultation.objects.filter(
        customer_id=user,
        design_id=portfolio
    ).order_by('-created_at').first()

    logger.debug(f"Latest consultation for user {user.id} and portfolio {portfolio_id}: {latest_consultation}")

    consultation_status = None
    if latest_consultation:
        consultation_status = latest_consultation.consultation_status
        logger.debug(f"Latest consultation status: {consultation_status}")

    if request.method == 'POST':
        if 'reject' in request.POST:
            # Handle rejection logic here
            cancelled_consultations = Consultation.objects.filter(
                customer_id=user,
                design_id=portfolio,
                consultation_status='Requested'
            ).update(consultation_status='Cancelled')
            logger.debug(f"Cancelled {cancelled_consultations} consultations")
            
            # Refresh the latest consultation after cancellation
            latest_consultation = Consultation.objects.filter(
                customer_id=user,
                design_id=portfolio,
                consultation_status__in=['Requested', 'Scheduled', 'Completed']
            ).order_by('-created_at').first()
            
            consultation_status = latest_consultation.consultation_status if latest_consultation else None
            logger.debug(f"Updated consultation status after cancellation: {consultation_status}")

    context = {
        'portfolio': portfolio,
        'user_type': user_type,
        'consultation_status': consultation_status,
        'user_details_complete': user_details_complete,
        'is_designer': request.user.is_authenticated and request.user == portfolio.designer_id,
    }
    return render(request, 'portfolio_details.html', context)
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from .models import ChatMessage, Design, Users


@login_required
@require_POST
def send_chat_message(request):
    data = json.loads(request.body)
    content = data.get('content')
    design_id = data.get('design_id')

    if content and design_id:
        design = Design.objects.get(id=design_id)
        sender = request.user
        
        # Determine the receiver based on the sender's user type
        if sender == design.designer_id:
            # If the sender is the designer, find the customer who initiated the consultation
            consultation = Consultation.objects.filter(design_id=design, designer_id=sender).first()
            receiver = consultation.customer_id if consultation else None
        else:
            # If the sender is not the designer (i.e., a customer), set the receiver to the designer
            receiver = design.designer_id

        message = ChatMessage.objects.create(
            sender=sender,
            receiver=receiver,
            design=design,
            content=content
        )
        return JsonResponse({'status': 'success', 'message_id': message.id})
    return JsonResponse({'status': 'error', 'message': 'Invalid data'})

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import models
from .models import Design, ChatMessage, Users
import json

@login_required
def get_chat_messages(request):
    design_id = request.GET.get('design_id')
    last_message_id = request.GET.get('last_message_id')
    
    if design_id:
        design = Design.objects.get(id=design_id)
        
        # Get messages newer than the last_message_id
        messages = ChatMessage.objects.filter(design=design, id__gt=last_message_id).order_by('timestamp')
        
        # Mark messages as read if the user is the receiver
        messages.filter(receiver=request.user, is_read=False).update(is_read=True)
        
        messages_data = [
            {
                'id': message.id,
                'sender_username': message.sender.username,
                'content': message.content,
                'timestamp': message.timestamp.isoformat(),
                'is_sender': message.sender == request.user,
                'is_designer': message.sender == design.designer_id
            }
            for message in messages
        ]
        return JsonResponse({
            'status': 'success', 
            'messages': messages_data,
            'designer_username': design.designer_id.username,
            'current_user': request.user.username
        })
    return JsonResponse({'status': 'error', 'message': 'Invalid design_id'})


@login_required
@require_POST
def clear_chat_history(request):
    """Clear chat history for a specific design between users"""
    data = json.loads(request.body)
    design_id = data.get('design_id')
    
    if not design_id:
        return JsonResponse({'status': 'error', 'message': 'Design ID is required'})
        
    try:
        design = Design.objects.get(id=design_id)
        
        # Clear messages where the user is either sender or receiver
        ChatMessage.objects.filter(
            design=design
        ).filter(
            Q(sender=request.user) | Q(receiver=request.user)
        ).delete()
        
        return JsonResponse({'status': 'success', 'message': 'Chat history cleared successfully'})
    except Design.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Design not found'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@login_required
@csrf_exempt
def consultation_booking(request, portfolio_id):
    portfolio = get_object_or_404(Design.objects.select_related('designer_id', 'amount'), id=portfolio_id)
    
    available_dates = ConsultationDate.objects.filter(
        designer=portfolio.designer_id,
        date_time__gte=timezone.now(),
        is_booked=False
    ).order_by('date_time')

    if request.method == 'POST':
        errors = {}
        try:
            schedule_date_time = request.POST.get('schedule_date')
            room_length = request.POST.get('room_length', '')
            room_width = request.POST.get('room_width', '')
            room_height = request.POST.get('room_height', '')
            design_preferences = request.POST.get('design_preferences')

            # Validate input data
            if not schedule_date_time:
                errors['schedule_date'] = 'Please select a consultation date.'
            
            for field, value in [('room_length', room_length), ('room_width', room_width), ('room_height', room_height)]:
                try:
                    decimal_value = Decimal(value)
                    if decimal_value <= 0:
                        errors[field] = f'{field.replace("_", " ").title()} must be a positive number.'
                except InvalidOperation:
                    errors[field] = f'Please enter a valid number for {field.replace("_", " ")}.'

            if not design_preferences:
                errors['design_preferences'] = 'Please enter your design preferences.'

            if errors:
                return JsonResponse({'status': 'error', 'errors': errors})

            # If we've made it this far, all inputs are valid
            consultation_amount = Amount.objects.create(amount=Decimal('500.00'))

            schedule_date_time = parse_datetime(schedule_date_time)
            if not schedule_date_time:
                return JsonResponse({'status': 'error', 'message': 'Invalid date-time format'})

            consultation_date = ConsultationDate.objects.filter(
                designer=portfolio.designer_id,
                date_time=schedule_date_time,
                is_booked=False
            ).first()

            if not consultation_date:
                return JsonResponse({'status': 'error', 'message': 'Selected date and time is not available'})

            consultation = Consultation(
                customer_id=request.user,
                designer_id=portfolio.designer_id,
                design_id=portfolio,
                date_time=consultation_date.date_time,
                consultation_status='Requested',
                proposal='Pending',
                schedule_date_time=consultation_date.date_time,
                room_length=Decimal(room_length),
                room_width=Decimal(room_width),
                room_height=Decimal(room_height),
                design_preferences=design_preferences,
                payment_type=Payment_Type.objects.get(payment_type='online'),
                payment_status='Paid',
                amount=consultation_amount,
                created_at=timezone.now()
            )
            consultation.save()

            consultation_date.is_booked = True
            consultation_date.save()

            return JsonResponse({'status': 'success', 'message': 'Consultation booked successfully'})
        
        except Exception as e:
            logger.error(f"Unexpected error in consultation booking: {str(e)}")
            return JsonResponse({'status': 'error', 'message': 'An unexpected error occurred. Please try again.'})

    context = {
        'portfolio': portfolio,
        'user_type': request.user.user_type_id.user_type if request.user.user_type_id else None,
        'available_dates': available_dates,
    }
    return render(request, 'consultation_booking.html', context)


from decimal import Decimal, InvalidOperation
from django.core.exceptions import ValidationError
from django.db.models import Avg
from django.utils import timezone
from datetime import timedelta

@nocache
def product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    related_products = Product.objects.filter(category=product.category).exclude(id=product.id)[:3]
    
    # Calculate delivery date (7 days from today)
    delivery_date = timezone.now().date() + timedelta(days=7)
    
    # Format the delivery date
    formatted_delivery_date = delivery_date.strftime("%A, %B %d, %Y")
    
    if request.method == 'POST' and request.user.is_authenticated:
        review_content = request.POST.get('review')
        rating = request.POST.get('rating')  # Add this line to get the rating
        if review_content and rating:
            Review.objects.create(
                product=product,
                user=request.user,
                content=review_content,
                rating=int(rating)  # Add this line to save the rating
            )
            messages.success(request, "Your review has been posted successfully.")
            return redirect('product', product_id=product.id)
        else:
            messages.error(request, "Review content and rating cannot be empty.")
    
    # Fetch reviews and count after potential new review creation
    reviews = Review.objects.filter(product=product).order_by('-created_at')
    review_count = reviews.count()
    
    # Calculate average rating
    avg_rating = reviews.aggregate(Avg('rating'))['rating__avg'] or 0
    
    context = {
        'product': product,
        'related_products': related_products,
        'reviews': reviews,
        'review_count': review_count,
        'avg_rating': round(avg_rating, 1),  # Round to 1 decimal place
        'product_amount_paise': int(product.amount.amount * 100),  # Convert to paise
        'csrf_token': get_token(request),  # Add CSRF token to context
        'delivery_date': formatted_delivery_date,  # Add this line
    }

    print(f"Debug - Delivery Date: {formatted_delivery_date}")  # Add this debug print

    if request.user.is_authenticated:
        user = request.user
        user_type = user.user_type_id.user_type if user.user_type_id else None
        context['user_type'] = user_type
        context['user'] = user

        # Get cart items for the current user
        cart_items = Cart.objects.filter(user_id=user).select_related('product_id__amount')
        context['cart_items'] = cart_items
        context['cart_total'] = cart_items.aggregate(total=Sum('amount'))['total'] or 0
        context['cart_count'] = cart_items.count()

        if request.method == 'POST':
            if 'product_id' in request.POST:
                product_id = int(request.POST.get('product_id'))
                status = request.POST.get('status')
                quantity = int(request.POST.get('quantity', 1))
                amount_value = request.POST.get('amount')
                
                if status == 'Added':
                    product_instance = get_object_or_404(Product, id=product_id)
                    
                    try:
                        unit_price = Decimal(amount_value)
                    except (InvalidOperation, TypeError):
                        unit_price = product_instance.amount.amount

                    cart_item, created = Cart.objects.get_or_create(
                        user_id=user,
                        product_id=product_instance,
                        defaults={'quantity': quantity, 'amount': unit_price * quantity, 'status': status}
                    )
                    
                    if not created:
                        cart_item.quantity += quantity
                        cart_item.amount = unit_price * cart_item.quantity
                        cart_item.save()

                    messages.success(request, 'Product added to cart successfully.')
                    return redirect('product', product_id=product_id)

    # Get related products (same category, excluding the current product)
    related_products = Product.objects.filter(category=product.category).exclude(id=product.id)[:4]
    context['related_products'] = related_products

    # Get and clear the cart message from the session
    cart_message = messages.get_messages(request)
    context['cart_message'] = cart_message

    return render(request, 'product.html', context)



@login_required
@never_cache
def cart(request):
    cart_items = Cart.objects.filter(user_id=request.user)
    
    # Calculate total price
    total_price = sum(item.amount for item in cart_items)

    context = {
        'cart_items': cart_items,
        'total_price': total_price,
    }
    return render(request, 'cart.html', context)

@login_required
@never_cache
@require_POST
@csrf_protect
def update_cart_quantity(request):
    import json
    from decimal import Decimal
    data = json.loads(request.body)
    item_id = data.get('item_id')
    quantity = int(data.get('quantity'))

    try:
        cart_item = get_object_or_404(Cart, id=item_id, user_id=request.user)
        product = cart_item.product_id

        # Calculate unit price if `amount` was previously calculated as total amount
        unit_price = cart_item.amount / cart_item.quantity
        cart_item.quantity = quantity
        cart_item.amount = unit_price * quantity
        
        # Save the updated cart item
        cart_item.save()

        # Recalculate total price for the updated cart
        cart_items = Cart.objects.filter(user_id=request.user)
        total_price = sum(item.amount for item in cart_items)

        # Check if any item in the cart is out of stock
        any_out_of_stock = any(item.quantity > item.product_id.stock for item in cart_items)

        return JsonResponse({
            'success': True,
            'total_price': float(total_price),
            'item_total': float(cart_item.amount),
            'total_items': sum(item.quantity for item in cart_items),
            'is_out_of_stock': quantity > product.stock,
            'available_stock': product.stock,
            'any_out_of_stock': any_out_of_stock
        })
    except Cart.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Cart item not found'}, status=400)

@login_required
@never_cache
def remove_from_cart(request, item_id):
    cart_item = get_object_or_404(Cart, id=item_id, user_id=request.user.id)
    cart_item.delete()
    
    # Recalculate total price and items count
    cart_items = Cart.objects.filter(user_id=request.user)
    total_price = sum(item.amount for item in cart_items)
    total_items = sum(item.quantity for item in cart_items)
    
    return JsonResponse({
        'success': True,
        'total_price': float(total_price),
        'total_items': total_items
    })
@login_required
@require_POST
def clear_cart(request):
    Cart.objects.filter(user_id=request.user).delete()
    return JsonResponse({'status': 'success'})


@login_required
@require_POST
def create_order_from_product(request):
    try:
        with transaction.atomic():
            product_id = request.POST.get('product_id')
            quantity = int(request.POST.get('quantity', 1))
            
            product = Product.objects.get(id=product_id)
            
            # Check if there's enough stock
            if product.stock < quantity:
                return JsonResponse({'status': 'error', 'message': f'Not enough stock for {product.name}'}, status=400)
            
            # Create or get the Amount instance
            amount_value = product.amount.amount * quantity
            amount, _ = Amount.objects.get_or_create(amount=amount_value)
            
            # Get or create the Payment_Type instance
            payment_type, _ = Payment_Type.objects.get_or_create(payment_type='online')
            
            # Create the order
            order = Order.objects.create(
                user=request.user,
                product=product,
                quantity=quantity,
                amount=amount,
                order_date=timezone.now(),
                order_status='Pending',
                delivery_date=timezone.now() + timedelta(days=7),
                payment_type=payment_type,
                payment_status='Paid'
            )
            
            # Update product stock
            product.stock -= quantity
            product.save()
            
            return JsonResponse({
                'status': 'success',
                'message': 'Order created successfully.',
                'order_id': order.id
            })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

import logging

logger = logging.getLogger(__name__)

@login_required
@require_POST
def create_order_from_cart(request):
    try:
        with transaction.atomic():
            cart_items = Cart.objects.filter(user_id=request.user)
            
            if not cart_items.exists():
                return JsonResponse({'status': 'error', 'message': 'Your cart is empty'}, status=400)
            
            orders_created = []
            
            for cart_item in cart_items:
                product = cart_item.product_id
                quantity = cart_item.quantity
                
                # Check if there's enough stock
                if product.stock < quantity:
                    if quantity == 1:
                        return JsonResponse({'status': 'error', 'message': f'Not enough stock for {product.name}, only {product.stock} available.'}, status=400)
                    else:
                        continue  # Skip this item if quantity is greater than available stock
                
                # Create or get the Amount instance
                amount_value = product.amount.amount * quantity
                amount, _ = Amount.objects.get_or_create(amount=amount_value)
                
                # Get or create the Payment_Type instance
                payment_type, _ = Payment_Type.objects.get_or_create(payment_type='online')
                
                # Create the order
                order = Order.objects.create(
                    user=request.user,
                    product=product,
                    quantity=quantity,
                    amount=amount,
                    order_date=timezone.now(),
                    order_status='Pending',
                    delivery_date=timezone.now() + timedelta(days=7),
                    payment_type=payment_type,
                    payment_status='Paid'
                )
                
                orders_created.append(order)
                
                # Update product stock
                product.stock -= quantity
                product.save()
                
                # Remove item from cart
                cart_item.delete()
            
            return JsonResponse({
                'status': 'success',
                'message': f'Successfully created {len(orders_created)} orders.',
                'order_ids': [order.id for order in orders_created]
            })
    except Exception as e:
        logger.error(f"Error processing order: {e}")  # Log the error for debugging
        return JsonResponse({'status': 'error', 'message': 'An error occurred while processing your order. Please try again later.'}, status=500)

@nocache
@login_required
def orders(request):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None
    
    # Fetch orders for the current user
    orders = Order.objects.filter(user=user).select_related('product', 'amount').order_by('-order_date')
    
    context = {
        'orders': orders,
        'user_type': user_type,
        'user': user
    }
    return render(request, 'my_orders.html', context)





import logging

logger = logging.getLogger(__name__)

@login_required
@require_http_methods(["GET", "POST"])
def schedule_consultation(request):
    if request.method == 'POST':
        date_time_str = request.POST.get('date_time')
        logger.info(f"Received date_time: {date_time_str}")
        try:
            # Parse the datetime string
            date_time = parse_datetime(date_time_str)
            logger.info(f"Parsed date_time: {date_time}")
            
            # Ensure the datetime is timezone aware
            if timezone.is_naive(date_time):
                date_time = timezone.make_aware(date_time, timezone.get_current_timezone())
            logger.info(f"Timezone-aware date_time: {date_time}")
            
            # Convert to the project's timezone if different
            date_time = timezone.localtime(date_time)
            logger.info(f"Localized date_time: {date_time}")

            if date_time < timezone.now():
                return JsonResponse({'status': 'error', 'message': 'Cannot schedule consultations for past dates'})

            consultation_date, created = ConsultationDate.objects.get_or_create(
                designer=request.user,
                date_time=date_time,
                defaults={'is_booked': False}
            )
            logger.info(f"Saved date_time: {consultation_date.date_time}")

            if created:
                return JsonResponse({'status': 'success', 'message': 'Consultation date scheduled successfully'})
            else:
                return JsonResponse({'status': 'error', 'message': 'This date and time is already scheduled'})

        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid date-time format'})

    else:
        # Handle GET request
        consultation_dates = ConsultationDate.objects.filter(designer=request.user).values('date_time', 'is_booked')
        consultation_dates_list = [
            {
                'id': date['date_time'].isoformat(),
                'title': 'Consultation',
                'start': date['date_time'].isoformat(),
                'allDay': False,
                'className': 'past-date' if date['date_time'] < timezone.now() else ('booked-date' if date['is_booked'] else ''),
                'editable': date['date_time'] >= timezone.now() and not date['is_booked'],
                'backgroundColor': '#FF0000' if date['is_booked'] else '#292929',  # Red for booked, default color otherwise
                'borderColor': '#FF0000' if date['is_booked'] else '#292929',
            } for date in consultation_dates
        ]
        
        context = {
            'consultation_dates': json.dumps(consultation_dates_list)
        }
        return render(request, 'schedule_consultation.html', context)


@login_required
@require_POST
def remove_scheduled_date(request):
    date_time_str = request.POST.get('date_time')
    try:
        date_time = parse_datetime(date_time_str)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Invalid date-time format'})

    if not date_time:
        return JsonResponse({'status': 'error', 'message': 'Invalid date-time format'})

    consultation_date = get_object_or_404(ConsultationDate, designer=request.user, date_time=date_time)

    if consultation_date.is_booked:
        return JsonResponse({'status': 'error', 'message': 'Cannot remove a booked consultation date'})

    consultation_date.delete()
    return JsonResponse({'status': 'success', 'message': 'Consultation date removed successfully'})


@nocache
@login_required
@require_POST
def edit_portfolio(request, portfolio_id):
    # Ensure the user is a designer and owns this portfolio
    portfolio = get_object_or_404(Design, id=portfolio_id, designer_id=request.user)
    
    if request.user.user_type_id.user_type != 'Designer':
        return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=403)

    name = request.POST.get('name')
    description = request.POST.get('description')
    image = request.FILES.get('image')
    amount_value = request.POST.get('amount')
    category = request.POST.get('category')  # New: Get the category from the POST request
    sqft = request.POST.get('sqft')  # New: Get the sqft from the POST request

    try:
        portfolio.name = name
        portfolio.description = description
        
        if image:
            portfolio.image = image
        
        if amount_value:
            # Update or create the amount
            if portfolio.amount:
                portfolio.amount.amount = amount_value
                portfolio.amount.save()
            else:
                amount = Amount.objects.create(amount=amount_value)
                portfolio.amount = amount

        if category:  # New: Update the category if provided
            portfolio.category = category

        if sqft:  # New: Update the sqft if provided
            portfolio.sqft = sqft

        portfolio.save()

        return redirect('portfolio_details', portfolio_id=portfolio_id)

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    

from django.core.mail import send_mail
from django.conf import settings

import logging
from django.core.mail import send_mail
from django.conf import settings
from django.contrib import messages

logger = logging.getLogger(__name__)

@nocache
@login_required
def consultations(request):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None
    designer = Users.objects.get(username=user.username)
    consultations = Consultation.objects.filter(designer_id=designer).select_related('customer_id', 'design_id').order_by('-id')

    for consultation in consultations:
        # Combine room dimensions into a single string
        consultation.room_dimensions = f"{consultation.room_length}L x {consultation.room_width}W x {consultation.room_height}H"
        
        # Localize and format the date_time
        consultation.schedule_date_time = timezone.localtime(consultation.schedule_date_time)
    
    if request.method == 'POST':
        if 'consultation_id' in request.POST:
            consultation_id = request.POST.get('consultation_id')
            action = request.POST.get('action')
            
            consultation = Consultation.objects.get(id=consultation_id)
            
            if action == 'approve':
                consultation.consultation_status = 'Scheduled'
                consultation.save()
                
                # Format the date and time
                formatted_datetime = consultation.schedule_date_time.astimezone(timezone.get_current_timezone())
                formatted_date = date_format(formatted_datetime, format="l, F j, Y")
                formatted_time = date_format(formatted_datetime, format="g:i A")

                # Recalculate room_dimensions here
                room_dimensions = f"{consultation.room_length}L x {consultation.room_width}W x {consultation.room_height}H"

                # Send email to the customer
                subject = 'Consultation Approved - ElegantDecor'
                message = f"""
                Dear {consultation.customer_id.name},

                Your consultation request has been approved and scheduled with ElegantDecor.

                Consultation Details:
                - Date: {formatted_date}
                - Time: {formatted_time}
                - Designer: {designer.name}
                - Designer's Contact Number: {designer.phone}
                - Design: {consultation.design_id.name}
                - Room Dimensions: {room_dimensions}
                - Your Design Preferences: {consultation.design_preferences}

                If you need to make any changes or have any questions, please don't hesitate to contact us.

                Thank you for choosing ElegantDecor!

                Best regards,
                The ElegantDecor Team
                """

                from_email = settings.DEFAULT_FROM_EMAIL
                recipient_list = [consultation.customer_id.email]

                logger.info(f"Attempting to send email to {recipient_list}")
                
                try:
                    send_mail(subject, message, from_email, recipient_list, fail_silently=False)
                    logger.info("Email sent successfully")
                except Exception as e:
                    logger.error(f"Error sending email: {str(e)}")
                    messages.error(request, "Consultation approved, but there was an error sending the email.")
                
            elif action == 'cancel':
                consultation.consultation_status = 'Pending'
                consultation.save()
            elif action == 'complete':
                consultation.consultation_status = 'Completed'
                consultation.save()
            
            return redirect('consultations')
    
    context = {
        'consultations': consultations,
        'user_type': user_type,
    }
    return render(request, 'consultations.html', context)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Consultation, Project
from django.views.decorators.http import require_http_methods
from django.views.decorators.cache import never_cache
from django.db import transaction


@never_cache
@login_required
@require_http_methods(["GET", "POST"])
def my_consultations(request):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None
    
    consultations = Consultation.objects.filter(customer_id=user).select_related('design_id', 'designer_id').order_by('-id')
    
    for consultation in consultations:
        # Format room dimensions
        consultation.room_dimensions = f"{consultation.room_length}L x {consultation.room_width}W x {consultation.room_height}H"
        
        # Localize and format the date_time
        consultation.schedule_date_time = timezone.localtime(consultation.schedule_date_time)
    
    if request.method == 'POST':
        consultation_id = request.POST.get('consultation_id')
        action = request.POST.get('action')
        
        if consultation_id and action:
            try:
                with transaction.atomic():
                    consultation = get_object_or_404(Consultation, id=consultation_id, customer_id=user)
                    
                    if consultation.consultation_status == 'Completed' and consultation.proposal == 'Pending':
                        if action == 'accept':
                            consultation.proposal = 'Accepted'
                            consultation.save()
                            
                            # Create a new project
                            Project.objects.create(
                                consultation=consultation,
                                design=consultation.design_id,
                                customer=consultation.customer_id,
                                designer=consultation.designer_id,
                                room_length=consultation.room_length,
                                room_width=consultation.room_width,
                                room_height=consultation.room_height
                            )
                            
                            messages.success(request, 'Proposal accepted and project created successfully!')
                        elif action == 'reject':
                            consultation.proposal = 'Rejected'
                            consultation.save()
                            messages.success(request, 'Proposal rejected successfully!')
                        else:
                            messages.error(request, 'Invalid action.')
                    else:
                        messages.error(request, 'Cannot update proposal for this consultation.')
            except Exception as e:
                messages.error(request, f'An error occurred: {str(e)}')
        else:
            messages.error(request, 'Invalid form submission.')
        
        return redirect('my_consultations')
    
    context = {
        'consultations': consultations,
        'user_type': user_type,
    }
    return render(request, 'my_consultations.html', context)



@login_required
def add_designer(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        
        # Validate username existence
        if Users.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return render(request, 'admin_page/add_designer.html')
        
        # Validate email existence
        if Users.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists.')
            return render(request, 'admin_page/add_designer.html')

        # Validate password criteria
        password_pattern = re.compile(r'^(?=.*[A-Z])(?=.*[a-z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$')
        if not password_pattern.match(password):
            messages.error(request, 'Password must be at least 8 characters long, contain at least one uppercase letter, one lowercase letter, one number, and one special character.')
            return render(request, 'admin_page/add_designer.html')
        
        # Check if passwords match
        if password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return render(request, 'admin_page/add_designer.html')
        
        try:
            designer_type = UserType.objects.get(user_type='Designer')
            
            new_designer = Users.objects.create(
                username=username,
                email=email,
                password=make_password(password),
                user_type_id=designer_type
            )
            new_designer.save()
            
            # Send email with username and password
            send_mail(
                'Welcome to ElegantDecor',
                f'Dear {username},\n\nYour account has been successfully created. Your login credentials are as follows:\nUsername: {username}\nPassword: {password}\n\nBest regards,\nThe ElegantDecor Team',
                'your-email@gmail.com',  # Replace with your actual email
                [email],
                fail_silently=False,
            )
            
            messages.success(request, 'Designer added successfully.')
            return redirect('tables')
        except Exception as e:
            messages.error(request, f'Error adding designer: {str(e)}')
    
    return render(request, 'admin_page/add_designer.html')



@login_required
def add_deliveryboy(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        
        # Validate username existence
        if Users.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return render(request, 'admin_page/add_deliveryboy.html')
        
        # Validate email existence
        if Users.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists.')
            return render(request, 'admin_page/add_deliveryboy.html')

        # Validate password criteria
        password_pattern = re.compile(r'^(?=.*[A-Z])(?=.*[a-z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$')
        if not password_pattern.match(password):
            messages.error(request, 'Password must be at least 8 characters long, contain at least one uppercase letter, one lowercase letter, one number, and one special character.')
            return render(request, 'admin_page/add_deliveryboy.html')
        
        # Check if passwords match
        if password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return render(request, 'admin_page/add_deliveryboy.html')
        
        try:
            # Get or create the Delivery_boy UserType
            deliveryboy_type, created = UserType.objects.get_or_create(
                user_type='Delivery_boy',
                defaults={'user_type': 'Delivery_boy'}
            )
            
            new_deliveryboy = Users.objects.create(
                username=username,
                email=email,
                password=make_password(password),
                user_type_id=deliveryboy_type
            )
            new_deliveryboy.save()
            
            # Send email with username and password
            send_mail(
                'Welcome to ElegantDecor',
                f'Dear {username},\n\nYour account has been successfully created. Your login credentials are as follows:\nUsername: {username}\nPassword: {password}\n\nBest regards,\nThe ElegantDecor Team',
                'your-email@gmail.com',  # Replace with your actual email
                [email],
                fail_silently=False,
            )
            
            messages.success(request, 'Delivery boy added successfully.')
            return redirect('tables')
        except Exception as e:
            messages.error(request, f'Error adding delivery boy: {str(e)}')
    
    return render(request, 'admin_page/add_deliveryboy.html')

@login_required
@never_cache
def deliveryboy_index(request):
    # Ensure the user is a delivery boy
    if not request.user.user_type_id or request.user.user_type_id.user_type != 'Delivery_boy':
        return redirect('signin')
        
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None
    
    context = {
        'user': user,
        'user_type': user_type,
    }
    return render(request, 'delivery_boy/index.html', context)

@login_required
@never_cache
def delivery(request):
    # Get the current delivery boy's orders
    my_orders = Order.objects.filter(
        delivery_boy=request.user,
        order_status='Approved'
    ).select_related(
        'user',  # For customer details
        'product',  # For product details
        'amount'  # For order amount
    ).order_by('-order_date')

    # Handle POST requests for updating delivery status
    if request.method == 'POST':
        order_id = request.POST.get('order_id')
        action = request.POST.get('action')
        
        try:
            order = Order.objects.get(id=order_id)
            
            if action == 'out_for_delivery':
                order.delivery_status = 'Out for Delivery'
                order.save()
                messages.success(request, 'Order marked as out for delivery')
                
            elif action == 'delivered':
                order.delivery_status = 'Delivered'
                order.delivery_date = timezone.now()
                order.save()
                messages.success(request, 'Order marked as delivered')
                
            elif action == 'accept':
                if not order.delivery_boy:
                    # Check if delivery boy is from same district
                    if order.user.district == request.user.district:
                        order.delivery_boy = request.user
                        order.delivery_status = 'Pending'
                        order.save()
                        messages.success(request, 'Order accepted successfully')
                    else:
                        messages.error(request, 'You can only accept orders from your district')
                else:
                    messages.error(request, 'This order is already assigned')
                    
        except Order.DoesNotExist:
            messages.error(request, 'Order not found')
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
            
        return redirect('delivery')

    # Get available orders in the delivery boy's district
    available_orders = Order.objects.filter(
        delivery_boy__isnull=True,
        order_status='Approved',
        user__district=request.user.district
    ).select_related(
        'user',
        'product',
        'amount'
    ).order_by('-order_date')

    # Combine my orders and available orders
    context = {
        'my_orders': my_orders,
        'available_orders': available_orders,
        'user': request.user,
        'delivery_stats': {
            'pending': my_orders.filter(delivery_status='Pending').count(),
            'out_for_delivery': my_orders.filter(delivery_status='Out for Delivery').count(),
            'delivered': my_orders.filter(delivery_status='Delivered').count(),
            'total': my_orders.count()
        }
    }
    
    return render(request, 'delivery_boy/delivery.html', context)

@login_required
@never_cache
def account(request):
    if request.method == 'POST':
        if 'status_update' in request.POST:
            try:
                new_status = request.POST.get('status').lower()  # Convert to lowercase
                if new_status in ['active', 'inactive']:
                    request.user.status = new_status
                    request.user.save()
                    messages.success(request, f'Your status has been updated to {new_status.capitalize()}')
                else:
                    messages.error(request, 'Invalid status value')
            except Exception as e:
                messages.error(request, 'Failed to update status')
            return redirect('account')

        # Existing profile update logic
        try:
            name = request.POST.get('name')
            email = request.POST.get('email')
            phone = request.POST.get('phone')
            home_town = request.POST.get('home_town')
            address = request.POST.get('address')
            district = request.POST.get('district')
            state = request.POST.get('state')
            pincode = request.POST.get('pincode')

            request.user.name = name
            request.user.email = email
            request.user.phone = phone
            request.user.home_town = home_town
            request.user.address = address
            request.user.district = district
            request.user.state = state
            request.user.pincode = pincode
            request.user.save()

            messages.success(request, 'Profile updated successfully')
        except Exception as e:
            messages.error(request, 'Failed to update profile')
        return redirect('account')

    context = {
        'user': request.user,
    }
    return render(request, 'delivery_boy/account.html', context)

@nocache
@login_required
def customers_table(request):
    customers = Users.objects.filter(user_type_id__user_type='Customer')
    designers = Users.objects.filter(user_type_id__user_type='Designer')

    if 'download' in request.GET:
        user_type = request.GET.get('user_type', '')
        if user_type not in ['customer', 'designer']:
            return HttpResponse('Invalid user type')

        if request.GET['download'] == 'pdf':
            return download_pdf(request, 'users', user_type)
        elif request.GET['download'] == 'excel':
            return download_excel(request, 'users', user_type)

    return render(request, 'admin_page/customers_table.html', {
        'customers': customers,
        'designers': designers,
    })

@nocache
@login_required
def designers_table(request):
    # Get download parameter from URL
    download_type = request.GET.get('download')
    user_type = 'designer'  # Specify user type for download

    if download_type:
        if download_type == 'excel':
            return download_excel(request, 'users', user_type)
        elif download_type == 'pdf':
            return download_pdf(request, 'users', user_type)

    # Get all designers
    designers = Users.objects.filter(user_type_id__user_type='Designer')
    
    return render(request, 'admin_page/designers_table.html', {  # Changed from users_table.html to designers_table.html
        'designers': designers,
    })

@nocache
@login_required
def designs_table(request):
    designs = Design.objects.select_related('designer_id', 'amount').all()
    
    if 'download' in request.GET:
        if request.GET['download'] == 'pdf':
            return download_pdf(request, 'designs')
        elif request.GET['download'] == 'excel':
            return download_excel(request, 'designs')
    
    context = {
        'designs': designs,
    }
    return render(request, 'admin_page/designs_table.html', context)

@nocache
@login_required
def projects_table(request):
    projects = Project.objects.select_related('customer', 'designer', 'design').all()
    
    if 'download' in request.GET:
        if request.GET['download'] == 'pdf':
            return download_pdf(request, 'projects')
        elif request.GET['download'] == 'excel':
            return download_excel(request, 'projects')
    
    context = {
        'projects': projects
    }
    return render(request, 'admin_page/projects_table.html', context)


from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from io import BytesIO

def download_pdf(request, data_type, user_type=None):
    template_path = 'admin_page/table_pdf_template.html'
    
    if data_type == 'users':
        if user_type == 'customer':
            filename = 'customers_table.pdf'
            title = 'Customers'
            users = Users.objects.filter(user_type_id__user_type='Customer')
        elif user_type == 'designer':
            filename = 'designers_table.pdf'
            title = 'Designers'
            users = Users.objects.filter(user_type_id__user_type='Designer')
        else:  # all users
            filename = 'all_users_table.pdf'
            title = 'All Users'
            users = Users.objects.exclude(user_type_id__user_type='Admin')

        headers = ['Name', 'Email', 'Phone', 'Address', 'Home Town', 'District', 'State', 'Pincode', 'Status', 'User Type']
        items = [
            [
                user.name,
                user.email,
                user.phone,
                user.address,
                user.home_town,
                user.district,
                user.state,
                user.pincode,
                user.status,
                user.user_type_id.user_type if user.user_type_id else 'N/A'
            ] for user in users
        ]
    elif data_type == 'designs':
        filename = 'designs_table.pdf'
        title = 'Designs'
        headers = ['Name', 'Description', 'Designer', 'Amount', 'Category', 'Square Feet']
        designs = Design.objects.select_related('designer_id', 'amount').all()
        items = [
            [
                design.name,
                design.description[:50] + ('...' if len(design.description) > 50 else ''),
                design.designer_id.name,
                str(design.amount.amount),
                design.category,
                str(design.sqft)
            ] for design in designs
        ]
    elif data_type == 'consultations':
        filename = 'consultations_table.pdf'
        title = 'Consultations'
        headers = ['Customer', 'Designer', 'Design', 'Status', 'Scheduled Date', 'Payment Type', 'Payment Status', 'Amount']
        consultations = Consultation.objects.select_related('customer_id', 'designer_id', 'design_id', 'payment_type', 'amount').all()
        items = [
            [
                consultation.customer_id.name,
                consultation.designer_id.name,
                consultation.design_id.name,
                consultation.consultation_status,
                str(consultation.schedule_date_time),
                consultation.payment_type.payment_type,
                consultation.payment_status,
                str(consultation.amount.amount)
            ] for consultation in consultations
        ]
    elif data_type == 'products':
        filename = 'products_table.pdf'
        title = 'Products'
        headers = ['Name', 'Description', 'Amount', 'Stock', 'Category']
        products = Product.objects.select_related('amount').all()
        items = [
            [
                product.name,
                product.description[:50] + ('...' if len(product.description) > 50 else ''),
                str(product.amount.amount),
                str(product.stock),
                product.category
            ] for product in products
        ]
    elif data_type == 'orders':
        filename = 'orders_table.pdf'
        title = 'Orders'
        headers = ['Order ID', 'User', 'Product', 'Quantity', 'Amount', 'Order Date', 'Order Status', 'Payment Type', 'Payment Status']
        orders = Order.objects.select_related('user', 'product', 'amount', 'payment_type').all()
        items = [
            [
                str(order.id),
                order.user.username,
                order.product.name,
                str(order.quantity),
                str(order.amount.amount),
                order.order_date.strftime('%Y-%m-%d %H:%M'),
                order.order_status,
                order.payment_type.payment_type,
                order.payment_status
            ] for order in orders
        ]
    elif data_type == 'projects':
        filename = 'projects_table.pdf'
        title = 'Projects'
        headers = ['Project ID', 'Customer', 'Designer', 'Design', 'Status', 'Start Date', 'Completed Date', 'Room Dimensions', 'Payment Status']
        projects = Project.objects.select_related('customer', 'designer', 'design').all()
        items = [
            [
                str(project.id),
                project.customer.name,
                project.designer.name,
                project.design.name,
                project.status,
                str(project.start_date) if project.start_date else 'N/A',
                str(project.completed_date) if project.completed_date else 'N/A',
                f"{project.room_length}x{project.room_width}x{project.room_height}",
                project.payment
            ] for project in projects
        ]
    else:
        return HttpResponse('Invalid data type')

    context = {
        'title': title,
        'headers': headers,
        'items': items,
    }
    # Create a Django response object, and specify content_type as pdf
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # find the template and render it.
    template = get_template(template_path)
    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(html, dest=response)
    # if error then show some funny view
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

def download_excel(request, data_type, user_type=None):
    workbook = Workbook()
    worksheet = workbook.active

    if data_type == 'users':
        if user_type == 'customer':
            worksheet.title = 'Customers'
            filename = 'customers_table.xlsx'
            data = Users.objects.filter(user_type_id__user_type='Customer')
        else:  # designer
            worksheet.title = 'Designers'
            filename = 'designers_table.xlsx'
            data = Users.objects.filter(user_type_id__user_type='Designer')

        headers = ['Name', 'Email', 'Phone', 'Address', 'Home Town', 'District', 'State', 'Pincode', 'Status']
    elif data_type == 'designs':
        worksheet.title = 'Designs'
        headers = ['Name', 'Description', 'Designer', 'Amount', 'Category', 'Square Feet']
        data = Design.objects.select_related('designer_id', 'amount').all()
        filename = 'designs_table.xlsx'
    elif data_type == 'consultations':
        worksheet.title = 'Consultations'
        headers = ['Customer', 'Designer', 'Design', 'Status', 'Scheduled Date', 'Payment Type', 'Payment Status', 'Amount']
        data = Consultation.objects.select_related('customer_id', 'designer_id', 'design_id').all()
        filename = 'consultations_table.xlsx'
    elif data_type == 'products':
        filename = 'products_table.xlsx'
        worksheet.title = 'Products'
        headers = ['Name', 'Description', 'Amount', 'Stock', 'Category']
        worksheet.append(headers)
        
        products = Product.objects.select_related('amount').all()
        for product in products:
            worksheet.append([
                product.name,
                product.description[:50] + ('...' if len(product.description) > 50 else ''),
                str(product.amount.amount),
                str(product.stock),
                product.category
            ])
        
        # Create a BytesIO buffer for the Excel file
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Create the HttpResponse object with Excel mime type
        response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'

        return response
    elif data_type == 'orders':
        filename = 'orders_table.xlsx'
        worksheet.title = 'Orders'
        headers = ['Order ID', 'User', 'Product', 'Quantity', 'Amount', 'Order Date', 'Order Status', 'Payment Type', 'Payment Status']
        worksheet.append(headers)
        
        orders = Order.objects.select_related('user', 'product', 'amount', 'payment_type').all()
        for order in orders:
            worksheet.append([
                str(order.id),
                order.user.username,
                order.product.name,
                str(order.quantity),
                str(order.amount.amount),
                order.order_date.strftime('%Y-%m-%d %H:%M'),
                order.order_status,
                order.payment_type.payment_type,
                order.payment_status
            ])
        
        # Create a BytesIO buffer for the Excel file
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Create the HttpResponse object with Excel mime type
        response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'

        return response
    elif data_type == 'projects':
        filename = 'projects_table.xlsx'
        worksheet.title = 'Projects'
        headers = ['Project ID', 'Customer', 'Designer', 'Design', 'Status', 'Start Date', 'Completed Date', 'Room Dimensions', 'Payment Status']
        worksheet.append(headers)
        
        projects = Project.objects.select_related('customer', 'designer', 'design').all()
        for project in projects:
            worksheet.append([
                str(project.id),
                project.customer.name,
                project.designer.name,
                project.design.name,
                project.status,
                str(project.start_date) if project.start_date else 'N/A',
                str(project.completed_date) if project.completed_date else 'N/A',
                f"{project.room_length}x{project.room_width}x{project.room_height}",
                project.payment
            ])
        
        # Create a BytesIO buffer for the Excel file
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Create the HttpResponse object with Excel mime type
        response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'

        return response
    else:
        return HttpResponse('Invalid data type')

    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header

    # Add data
    for row_num, item in enumerate(data, 2):
        if data_type == 'users':
            row = [
                item.name,
                item.email,
                item.phone,
                item.address,
                item.home_town,
                item.district,
                item.state,
                item.pincode,
                item.status,
            ]
        elif data_type == 'designs':
            row = [
                item.name,
                item.description[:50] + '...' if len(item.description) > 50 else item.description,
                item.designer_id.name,
                str(item.amount.amount),
                item.category,
                str(item.sqft)
            ]
        elif data_type == 'consultations':
            row = [
                item.customer_id.name,
                item.designer_id.name,
                item.design_id.name,
                item.consultation_status,
                str(item.schedule_date_time),
                item.payment_type.payment_type,
                item.payment_status,
                str(item.amount.amount)
            ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Create a BytesIO buffer for the Excel file
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Create the HttpResponse object with Excel mime type
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response

    
@nocache
@login_required
def consultations_table(request):
    consultations = Consultation.objects.select_related('customer_id', 'designer_id', 'design_id').all()
    
    if 'download' in request.GET:
        if request.GET['download'] == 'pdf':
            return download_pdf(request, 'consultations')
        elif request.GET['download'] == 'excel':
            return download_excel(request, 'consultations')
    
    context = {
        'consultations': consultations
    }
    return render(request, 'admin_page/consultations_table.html', context)

@nocache
@login_required
def orders_table(request):
    orders = Order.objects.all().order_by('-order_date')
    
    if request.method == 'POST':
        order_id = request.POST.get('order_id')
        action = request.POST.get('action')
        
        try:
            order = Order.objects.get(id=order_id)
            
            if action == 'approve':
                order.order_status = 'Approved'
                order.save()
                messages.success(request, 'Order approved successfully.')
            
            elif action == 'cancel':
                order.order_status = 'Cancelled'
                order.save()
                messages.success(request, 'Order cancelled successfully.')
            
            elif action == 'assign_delivery_boy':
                delivery_boy_id = request.POST.get('delivery_boy_id')
                if delivery_boy_id:
                    delivery_boy = Users.objects.get(id=delivery_boy_id)
                    # Verify delivery boy is from same district
                    if delivery_boy.district == order.user.district:
                        order.delivery_boy = delivery_boy
                        order.delivery_status = 'Pending'
                        order.save()
                        
                        messages.success(
                            request, 
                            f'Successfully assigned delivery boy: {delivery_boy.name}\n'
                            f'Contact: {delivery_boy.phone}\n'
                            f'District: {delivery_boy.district}'
                        )
                    else:
                        messages.error(request, 'Delivery boy must be from the same district as the customer.')
                else:
                    messages.error(request, 'Please select a delivery boy.')
                    
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
        
        return redirect('orders_table')

    # For each order, get delivery boys from the same district
    for order in orders:
        if order.order_status == 'Approved' and not order.delivery_boy:
            order.available_delivery_boys = Users.objects.filter(
                user_type_id__user_type='Delivery_boy',
                status='active',
                district=order.user.district
            ).select_related('user_type_id')

    context = {
        'orders': orders
    }
    
    return render(request, 'admin_page/orders_table.html', context)

@nocache
@login_required
@require_http_methods(["GET", "POST"])
def products_table(request):
    if request.method == 'POST':
        if '_method' in request.POST and request.POST['_method'] == 'DELETE':
            # Handle product delete
            product_id = request.POST.get('id')
            product = get_object_or_404(Product, id=product_id)
            product.delete()
            return JsonResponse({'status': 'success'})
        
        # Handle product update
        product_id = request.POST.get('id')
        product = get_object_or_404(Product, id=product_id)
        product.name = request.POST.get('name')
        product.description = request.POST.get('description')
        product.amount.amount = request.POST.get('amount')
        product.stock = request.POST.get('stock')
        product.category = request.POST.get('category')  # Make sure this line is present
        product.color = request.POST.get('color')
        product.save()
        product.amount.save()
        return JsonResponse({'status': 'success'})

    # For GET requests, handle downloads and render the table
    if 'download' in request.GET:
        if request.GET['download'] == 'pdf':
            return download_pdf(request, 'products')
        elif request.GET['download'] == 'excel':
            return download_excel(request, 'products')

    # For GET requests, render the table
    products = Product.objects.select_related('amount').all()
    context = {
        'products': products
    }
    return render(request, 'admin_page/products_table.html', context)

@nocache
@login_required
def admin_index(request):
    
    return render(request, 'admin_page/admin_index.html')


from django.utils import timezone
from datetime import timedelta

@nocache
@login_required
def orders(request):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None
    
    # Fetch orders for the current user
    orders = Order.objects.filter(user=user).select_related(
        'product', 'amount', 'payment_type', 'user'
    ).order_by('-order_date')
    
    # Calculate and add delivery date for each order
    for order in orders:
        order.delivery_date = order.order_date + timedelta(days=7)
    
    context = {
        'orders': orders,
        'user_type': user_type,
        'user': user
    }
    return render(request, 'my_orders.html', context)


from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO

@login_required
def download_receipt(request, order_id):
    try:
        # Get the order and verify ownership
        order = Order.objects.select_related(
            'user', 
            'product', 
            'payment_type',
            'amount'
        ).get(id=order_id)

        # Security check: Ensure user owns this order or is staff
        if order.user != request.user and not request.user.is_staff:
            return HttpResponseForbidden("You don't have permission to access this receipt.")

        # Verify order status
        if not (order.order_status == 'Approved' and order.delivery_status == 'Delivered'):
            return HttpResponseForbidden("Receipt is only available for delivered orders.")

        # Generate PDF
        template = get_template('receipt.html')
        context = {
            'order': order,
            'generated_date': timezone.now(),
        }
        html = template.render(context)

        # Create PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="order_receipt_{order_id}.pdf"'

        # Convert HTML to PDF
        pisa_status = pisa.CreatePDF(
            html, 
            dest=response,
            encoding='utf-8'
        )

        # Return PDF if successful
        if pisa_status.err:
            return HttpResponse('Error generating PDF', status=500)
        return response

    except Order.DoesNotExist:
        return HttpResponse('Order not found', status=404)
    except Exception as e:
        print(f"Error generating receipt: {str(e)}")
        return HttpResponse('Error generating receipt', status=500)

import cv2
import numpy as np
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import os
from django.conf import settings
from django.db.models import Q
import pandas as pd
from sklearn.neighbors import KNeighborsClassifier

# Load the dataset and train the KNN model
data = pd.read_csv(os.path.join(settings.BASE_DIR, 'data/colors.csv'))
X = data[['R', 'G', 'B']]
y = data['ColorName1']  # Use ColorName1 for the color names
knn = KNeighborsClassifier(n_neighbors=3)
knn.fit(X, y)

def detect_color(image_path):
    # Check if the file exists
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"The file {image_path} does not exist.")

    # Load the image using OpenCV
    image = cv2.imread(image_path)
    
    # Check if the image is loaded properly
    if image is None:
        raise ValueError(f"Failed to load image from {image_path}. The file may not be a valid image format.")
    
    # Resize image for faster processing
    image = cv2.resize(image, (300, 300))
    # Convert image from BGR to RGB
    image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    # Reshape image to be a list of pixels
    pixels = np.float32(image.reshape(-1, 3))

    # Use k-means clustering to find the dominant color
    n_colors = 1
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 200, 0.1)
    _, labels, palette = cv2.kmeans(pixels, n_colors, None, criteria, 10, cv2.KMEANS_RANDOM_CENTERS)

    # Convert the dominant color to a tuple
    dominant_color = palette[0].astype(int)
    
    # Predict the closest color using the KNN model
    detected_color = knn.predict([dominant_color])[0]
    return detected_color

def recommend_products_by_color(request):
    context = {}
    if request.user.is_authenticated:
        user_type = request.user.user_type_id.user_type if request.user.user_type_id else None
        context['user_type'] = user_type

    if request.method == 'POST' and 'image' in request.FILES:
        image = request.FILES['image']
        
        # Save the file and get the path
        file_name = default_storage.save(f'uploads/{image.name}', ContentFile(image.read()))
        file_path = os.path.join(settings.MEDIA_ROOT, file_name)
        
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"The file {file_path} does not exist after saving.")
            
            detected_color = detect_color(file_path)
            
            # Remove underscore and capitalize each word
            formatted_color = detected_color.replace('_', ' ').title()
            
            # Filter products based on the detected color
            color_query = Q(color__iexact=detected_color)
            recommended_products = Product.objects.filter(color_query)
            
            context.update({
                'detected_color': formatted_color,  # Use the formatted color name
                'recommended_products': recommended_products,
                'products_count': recommended_products.count(),
                'form_submitted': True,
                'image_url': default_storage.url(file_name)
            })
        except Exception as e:
            context.update({
                'error': str(e),
                'form_submitted': True
            })
    else:
        context['form_submitted'] = False
    
    return render(request, 'recommend_products.html', context)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from .models import MoodBoard, MoodBoardItem
@nocache
@login_required
def mood_board_list(request):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None

    mood_boards = MoodBoard.objects.filter(user=user)

    context = {
        'mood_boards': mood_boards,
        'user_type': user_type,
        'user': user,
        'no_results': mood_boards.count() == 0,
    }

    return render(request, 'mood_boards/list.html', context)

@login_required
def create_mood_board(request):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        image = request.FILES.get('image')
        MoodBoard.objects.create(user=user, name=name, description=description, image=image)
        return redirect('mood_board_list')
    return render(request, 'mood_boards/create.html', {'user_type': user_type})

@login_required
def mood_board_detail(request, pk):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None
    mood_board = get_object_or_404(MoodBoard, pk=pk, user=user)
    items = mood_board.items.all()
    return render(request, 'mood_boards/detail.html', {'mood_board': mood_board, 'items': items, 'user_type': user_type})

@login_required
def add_design_to_mood_board(request, pk):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None
    mood_board = get_object_or_404(MoodBoard, pk=pk, user=user)
    
    # Get filter parameters
    category = request.GET.get('category', 'all')
    designer = request.GET.get('designer', 'all')
    sqft_range = request.GET.get('sqft', 'all')
    search_query = request.GET.get('search', '')

    designs = Design.objects.select_related('designer_id', 'amount').all()

    if category and category != 'all':
        designs = designs.filter(category__iexact=category)

    if designer and designer != 'all':
        designs = designs.filter(designer_id__username=designer)

    if sqft_range and sqft_range != 'all':
        sqft_ranges = {
            '0-500': (0, 500),
            '501-1000': (501, 1000),
            '1001-1500': (1001, 1500),
            '1501+': (1501, float('inf'))
        }
        min_sqft, max_sqft = sqft_ranges.get(sqft_range, (None, None))
        if min_sqft is not None:
            designs = designs.filter(sqft__gte=min_sqft)
        if max_sqft != float('inf'):
            designs = designs.filter(sqft__lte=max_sqft)

    if search_query:
        designs = designs.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(designer_id__username__icontains=search_query)
        )

    all_designers = Users.objects.filter(user_type_id__user_type='Designer').values_list('username', flat=True).distinct()
    all_categories = Design.objects.values_list('category', flat=True).distinct()

    def format_category(category):
        return category.replace('_', ' ').title()

    formatted_categories = [{'value': cat, 'display': format_category(cat)} for cat in all_categories]

    context = {
        'mood_board': mood_board,
        'designs': designs,
        'user_type': user_type,
        'designers': all_designers,
        'categories': formatted_categories,
        'no_results': designs.count() == 0,
        'selected_category': category,
        'selected_designer': designer,
        'selected_sqft_range': sqft_range,
        'search_query': search_query,
    }

    return render(request, 'mood_boards/add_design.html', context)

@login_required
def add_product_to_mood_board(request, pk):
    mood_board = get_object_or_404(MoodBoard, pk=pk, user=request.user)
    products = Product.objects.all()
    
    # Get filter parameters
    category = request.GET.get('category', 'all')
    search_query = request.GET.get('search', '')
    
    # Filter by category
    if category and category != 'all':
        products = products.filter(category=category)
    
    # Filter by search query
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Get unique categories for the dropdown
    categories = Product.objects.values_list('category', flat=True).distinct()
    
    context = {
        'mood_board': mood_board,
        'products': products,
        'categories': categories,
        'selected_category': category,
        'search_query': search_query,
        'no_results': products.count() == 0,
    }

    if request.user.is_authenticated:
        user = request.user
        user_type = user.user_type_id.user_type if user.user_type_id else None
        context['user_type'] = user_type
        context['user'] = user

    return render(request, 'mood_boards/add_product.html', context)

@login_required
@require_POST
def add_mood_board_item(request, pk):
    # This function doesn't render a template, so we don't need to pass user_type here
    mood_board = get_object_or_404(MoodBoard, pk=pk, user=request.user)
    item_type = request.POST.get('item_type')
    item_id = request.POST.get('item_id')

    if item_type == 'design':
        design = get_object_or_404(Design, pk=item_id)
        item = MoodBoardItem.objects.create(
            mood_board=mood_board,
            item_type='design',
            design=design,
            position_x=0,
            position_y=0
        )
    elif item_type == 'product':
        product = get_object_or_404(Product, pk=item_id)
        item = MoodBoardItem.objects.create(
            mood_board=mood_board,
            item_type='product',
            product=product,
            position_x=0,
            position_y=0
        )
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid item type'})

    return JsonResponse({
        'status': 'success',
        'item_id': item.id,
        'image_url': item.image_url,
        'caption': item.caption,
        'position_x': item.position_x,
        'position_y': item.position_y
    })

@login_required
@require_POST
def update_item_position(request, pk):
    # This function doesn't render a template, so we don't need to pass user_type here
    item = get_object_or_404(MoodBoardItem, pk=pk, mood_board__user=request.user)
    position_x = request.POST.get('position_x')
    position_y = request.POST.get('position_y')
    
    if position_x is not None and position_y is not None:
        try:
            item.position_x = int(float(position_x))
            item.position_y = int(float(position_y))
            item.save()
            return JsonResponse({'status': 'success'})
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid position data'}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Missing position data'}, status=400)

from django.http import JsonResponse
from django.views.decorators.http import require_POST

@require_POST
def delete_mood_board_item(request, pk):
    try:
        item = MoodBoardItem.objects.get(pk=pk)
        item.delete()
        return JsonResponse({'status': 'success'})
    except MoodBoardItem.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Item not found'}, status=404)
    
    
@login_required
@require_POST
def delete_mood_board(request, pk):
    try:
        mood_board = MoodBoard.objects.get(pk=pk, user=request.user)
        mood_board.delete()
        return JsonResponse({'status': 'success'})
    except MoodBoard.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Mood board not found'}, status=404)
    

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Project

@login_required
def my_projects(request):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None
    
    if user_type == 'Customer':
        projects = Project.objects.filter(customer=user).select_related('design', 'designer')
    elif user_type == 'Designer':
        projects = Project.objects.filter(designer=user).select_related('design', 'customer')
    else:
        projects = []

    context = {
        'projects': projects,
        'user_type': user_type,
    }
    return render(request, 'my_projects.html', context)


from django.views.decorators.http import require_http_methods
from django.http import JsonResponse

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from .models import Project, ProjectFeedback

@login_required
@require_http_methods(["GET", "POST"])
def project(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    user_type = request.user.user_type_id.user_type if request.user.user_type_id else None
    
    # Check if feedback already exists
    existing_feedback = ProjectFeedback.objects.filter(project=project, customer=request.user).first()
    
    if request.method == 'POST':
        if 'payment_id' in request.POST:
            payment_id = request.POST.get('payment_id')
            if payment_id:
                # Update project payment status
                project.payment = 'paid'
                project.save()
                return JsonResponse({'status': 'success', 'user_type': user_type})
            else:
                return JsonResponse({'status': 'error', 'message': 'Invalid payment data', 'user_type': user_type})
        elif 'feedback' in request.POST:
            feedback = request.POST.get('feedback')
            if feedback:
                # Create new feedback or update existing
                if existing_feedback:
                    existing_feedback.feedback = feedback
                    existing_feedback.save()
                else:
                    existing_feedback = ProjectFeedback.objects.create(
                        project=project,
                        customer=request.user,
                        feedback=feedback
                    )
                return JsonResponse({'status': 'success', 'message': 'Feedback submitted successfully'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Invalid feedback data'})
    
    context = {
        'project': project,
        'user_type': user_type,
        'existing_feedback': existing_feedback,
    }
    return render(request, 'project.html', context)

@login_required
@require_http_methods(["GET", "POST"])
def projects_manage(request):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None
    
    if user_type != 'Designer':
        return redirect('index')  # Redirect non-designers to home page
    
    projects = Project.objects.filter(designer=user).select_related('design', 'customer').prefetch_related('feedbacks').order_by('-id')
    
    if request.method == 'POST':
        project_id = request.POST.get('project_id')
        action = request.POST.get('action')
        
        try:
            project = projects.get(id=project_id)
            
            if action == 'update_status':
                new_status = request.POST.get('status')
                if new_status in ['Not Started', 'In Progress', 'Completed', 'On Hold']:
                    project.status = new_status
                    if new_status == 'Completed' and not project.completed_date:
                        project.completed_date = timezone.now().date()
                    project.save()
                    return JsonResponse({
                        'status': 'success',
                        'message': 'Project status updated successfully',
                        'completed_date': project.completed_date.strftime('%b. %d, %Y') if project.completed_date else None
                    })
            elif action == 'update_start_date':
                start_date = request.POST.get('start_date')
                try:
                    project.start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                    project.status = 'In Progress'  # Automatically set status to 'In Progress' when start date is set
                    project.save()
                    return JsonResponse({'status': 'success', 'message': 'Start date updated successfully'})
                except ValueError:
                    return JsonResponse({'status': 'error', 'message': 'Invalid date format'})
            
            return JsonResponse({'status': 'error', 'message': 'Invalid action'})
        
        except Project.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Project not found'}, status=404)
    
    context = {
        'projects': projects,
        'user_type': user_type,
    }
    return render(request, 'projects_manage.html', context)



# Initialize Gemini with API key directly
# GEMINI_API_KEY ="AIzaSyDS4jx4S-2cgUnBM6Ouc5D6Xwzj-Dorf-E"
# GEMINI_API_KEY ="AIzaSyCcldFUfZNm4jRcDeWhj64mUfvqx7l3t18"

@login_required
def get_user_info(request):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None
    
    user_data = {
        'name': user.name,
        'email': user.email,
        'phone': user.phone,
        'user_type': user_type,
        'address': user.address,
        'home_town': user.home_town,
        'district': user.district,
        'state': user.state,
        'pincode': user.pincode,
    }

    if user_type in ['Designer', 'Customer']:
        # Get projects based on user type
        if user_type == 'Designer':
            projects = Project.objects.filter(designer=user).select_related(
                'design', 'customer'
            )
        else:  # Customer
            projects = Project.objects.filter(customer=user).select_related(
                'design', 'designer'
            )

        # Format projects data
        projects_data = [{
            'design': {
                'name': project.design.name if project.design else 'N/A'
            },
            'customer': {
                'name': project.customer.name if project.customer else 'N/A'
            },
            'designer': {
                'name': project.designer.name if project.designer else 'N/A'
            },
            'room_length': float(project.room_length),
            'room_width': float(project.room_width),
            'room_height': float(project.room_height),
            'status': project.status,
            'start_date': project.start_date.strftime('%Y-%m-%d') if project.start_date else None,
            'completed_date': project.completed_date.strftime('%Y-%m-%d') if project.completed_date else None
        } for project in projects]

        # Add projects data to user_data
        user_data.update({
            'projects': projects_data,
            'total_projects': len(projects_data),
            'active_projects': sum(1 for p in projects_data if p['status'] in ['Not Started', 'In Progress'])
        })

        # Get consultations with correct filter syntax
        if user_type == 'Customer':
            consultations = Consultation.objects.filter(customer_id=user).select_related(
                'design_id', 'customer_id', 'designer_id', 'amount'
            )
        else:  # Designer
            consultations = Consultation.objects.filter(designer_id=user).select_related(
                'design_id', 'customer_id', 'designer_id', 'amount'
            )

        # Format consultations data
        consultations_data = [{
            'design_name': cons.design_id.name,
            'customer_name': cons.customer_id.name,
            'customer_email': cons.customer_id.email,
            'customer_phone': cons.customer_id.phone,
            'designer_name': cons.designer_id.name,
            'designer_email': cons.designer_id.email,
            'designer_phone': cons.designer_id.phone,
            'status': cons.consultation_status,
            'schedule_date': cons.schedule_date_time.strftime('%Y-%m-%d %H:%M') if cons.schedule_date_time else 'Not Scheduled',
            'room_dimensions': f"{cons.room_length}x{cons.room_width}x{cons.room_height}",
            'payment_status': cons.payment_status,
            'amount': str(cons.amount.amount) if cons.amount else 'N/A'
        } for cons in consultations]

        user_data.update({
            'consultations': consultations_data,
            'total_consultations': len(consultations_data),
            'pending_consultations': sum(1 for c in consultations_data if c['status'] == 'Pending')
        })

    if user_type == 'Customer':
        # Get orders for customer
        orders = Order.objects.filter(user=user).select_related(
            'product', 'amount', 'payment_type'
        ).order_by('-order_date')

        # Format orders data with image
        orders_data = [{
            'order_id': order.id,
            'product': order.product.name,
            'product_image': order.product.image.url if order.product.image else None,
            'quantity': order.quantity,
            'amount': str(order.amount.amount),
            'order_date': order.order_date.strftime('%Y-%m-%d %H:%M'),
            'order_status': order.order_status,
            'payment_type': order.payment_type.payment_type if order.payment_type else 'N/A'
        } for order in orders]

        # Add orders data to user_data
        user_data['orders'] = orders_data
        user_data['total_orders'] = len(orders_data)
        user_data['pending_orders'] = sum(1 for o in orders_data if o['order_status'] == 'Pending')

    # Add designs data for designers
    if user_type == 'Designer':
        designs = Design.objects.filter(designer_id=user).select_related('amount')
        designs_data = [{
            'name': design.name,
            'description': design.description,
            'category': design.category,
            'sqft': float(design.sqft),
            'amount': float(design.amount.amount),
            'image_url': design.image.url if design.image else None
        } for design in designs]

        user_data.update({
            'designs': designs_data,
            'total_designs': len(designs_data)
        })

    return JsonResponse(user_data)

@nocache
@login_required
def product_sales_analytics(request):
    # Get all orders
    orders = Order.objects.select_related('product', 'amount').all()
    
    # Calculate total sales and revenue
    total_sales = orders.count()
    total_revenue = float(sum(order.amount.amount for order in orders))  # Convert to float
    # Get sales by product
    product_sales = {}
    for order in orders:
        product_name = order.product.name
        if product_name not in product_sales:
            product_sales[product_name] = {
                'count': 0,
                'revenue': 0,
                'image_url': order.product.image.url if order.product.image else None
            }
        product_sales[product_name]['count'] += order.quantity
        product_sales[product_name]['revenue'] += float(order.amount.amount)  # Convert to float

    # Calculate percentages and prepare data for charts
    top_products = sorted(
        [
            {
                'name': name,
                'count': data['count'],
                'revenue': data['revenue'],
                'image_url': data['image_url'],
                'percentage': float((data['count'] / total_sales * 100) if total_sales > 0 else 0)  # Convert to float
            }
            for name, data in product_sales.items()
        ],
        key=lambda x: x['revenue'],
        reverse=True
    )[:5]  # Get top 5 products

    # Get monthly sales data for the last 12 months
    monthly_sales = {}
    today = timezone.now()
    for i in range(12):
        month_start = today - timedelta(days=today.day - 1) - timedelta(days=30 * i)
        month_name = month_start.strftime('%B %Y')
        monthly_sales[month_name] = {
            'count': 0,
            'revenue': 0
        }

    for order in orders:
        month_name = order.order_date.strftime('%B %Y')
        if month_name in monthly_sales:
            monthly_sales[month_name]['count'] += order.quantity
            monthly_sales[month_name]['revenue'] += float(order.amount.amount)  # Convert to float

    # Convert monthly data to lists for Chart.js
    monthly_labels = list(reversed(list(monthly_sales.keys())))
    monthly_revenue = [float(monthly_sales[month]['revenue']) for month in monthly_labels]  # Convert to float

    context = {
        'total_sales': total_sales,
        'total_revenue': total_revenue,
        'top_products': top_products,
        'monthly_labels': json.dumps(monthly_labels),
        'monthly_revenue': json.dumps(monthly_revenue),
    }
    
    return render(request, 'admin_page/product_sales_analytics.html', context)

    
@login_required
def get_room_models(request):
    try:
        models = RoomModel.objects.all()
        models_data = [{
            'id': model.id,
            'name': model.name,
            'model_file': model.model_file.url if model.model_file else None,
            'thumbnail': model.thumbnail.url if model.thumbnail else None,
            'category': model.category,
            'created_at': model.created_at.isoformat()
        } for model in models]
        
        return JsonResponse({
            'models': models_data
        })
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)



@login_required
def add_models(request):
    # Get all models and handle the case where thumbnail might be None
    room_models = RoomModel.objects.all().order_by('-created_at')
    context = {
        'room_models': room_models,
        'user_type': 'admin'
    }
    return render(request, 'admin_page/add_models.html', context)

@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def delete_room_model(request, model_id):
    try:
        model = RoomModel.objects.get(id=model_id)
        
        # Safely delete files if they exist
        try:
            if model.model_file:
                model.model_file.delete(save=False)
        except:
            pass
            
        try:
            if model.thumbnail:
                model.thumbnail.delete(save=False)
        except:
            pass
            
        # Delete the model instance
        model.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Model deleted successfully'
        })
    except RoomModel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Model not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_http_methods(["POST"])
def upload_room_model(request):
    try:
        name = request.POST.get('name')
        category = request.POST.get('category')
        model_file = request.FILES.get('model_file')
        thumbnail = request.FILES.get('thumbnail')

        if not all([name, category, model_file, thumbnail]):
            return JsonResponse({'success': False, 'error': 'All fields are required'})

        # Validate file type
        if not model_file.name.endswith(('.gltf', '.glb')):
            return JsonResponse({'success': False, 'error': 'Invalid model file format'})

        # Create new room model - remove the user parameter
        room_model = RoomModel.objects.create(
            name=name,
            category=category,
            model_file=model_file,
            thumbnail=thumbnail
            # Remove the user field that's causing the error
        )

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

#frontend


import json
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.core.files.storage import default_storage
from .models import RoomModel

@login_required
def virtual_room_designer(request):
    user = request.user
    user_type = user.user_type_id.user_type if user.user_type_id else None
    
    rooms = VirtualRoom.objects.filter(user=request.user)
    
    context = {
        'rooms': rooms,
        'user_type': user_type,
    }
    return render(request, 'virtual_room/designer.html', context)

@login_required
def get_models(request):
    try:
        # Get all models from the database
        models = RoomModel.objects.all()
        
        # Prepare data to return
        models_data = []
        for model in models:
            model_data = {
                'id': model.id,
                'name': model.name,
                'category': model.category,
                'model_file': model.model_file.url if model.model_file else None,
                'thumbnail': model.thumbnail.url if model.thumbnail else None
            }
            models_data.append(model_data)
        
        return JsonResponse({
            'success': True,
            'models': models_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

@login_required
def get_rooms(request, room_id=None):
    """
    Combined endpoint to:
    1. Get all rooms for the current user (when room_id is None)
    2. Get a specific room with its models (when room_id is provided)
    """
    try:
        if room_id is None:
            # Get all rooms for the current user
            rooms = VirtualRoom.objects.filter(user=request.user).order_by('-updated_at')
            
            # Format the rooms as a list of dictionaries
            rooms_data = []
            for room in rooms:
                # Count models in this room
                model_count = VirtualRoomModel.objects.filter(virtual_room=room).count()
                
                rooms_data.append({
                    'id': room.id,
                    'name': room.name,
                    'width': float(room.width),
                    'length': float(room.length),
                    'height': float(room.height),
                    'model_count': model_count,
                    'wall_color': room.wall_color,
                    'floor_color': room.floor_color,
                    'ceiling_color': room.ceiling_color,
                    'created_at': room.created_at.strftime('%Y-%m-%d %H:%M'),
                    'updated_at': room.updated_at.strftime('%Y-%m-%d %H:%M')
                })
            
            return JsonResponse({
                'success': True,
                'rooms': rooms_data
            })
        else:
            # Get the specific room with all its models
            room = VirtualRoom.objects.get(id=room_id, user=request.user)
            
            # Get all virtual room models with their associated room models (3D models)
            virtual_room_models = VirtualRoomModel.objects.filter(virtual_room=room)

            room_data = {
                'success': True,
                'room': {
                    'id': room.id,
                    'name': room.name,
                    'width': float(room.width),
                    'length': float(room.length),
                    'height': float(room.height),
                    'wall_color': room.wall_color,
                    'floor_color': room.floor_color,
                    'ceiling_color': room.ceiling_color
                },
                'models': []
            }
            
            # Process each model in the room
            for vrm in virtual_room_models:
                model = vrm.room_model
                
                if model and model.model_file:  # Make sure model and its file exist
                    try:
                        model_url = model.model_file.url
                        absolute_url = request.build_absolute_uri(model_url)
                        
                        # Get category or infer it from name
                        category = model.category or get_category_from_name(model.name)
                        
                        model_data = {
                            'id': vrm.id,  # VirtualRoomModel ID
                            'model_id': model.id,  # RoomModel ID
                            'name': model.name,
                            'category': category,
                            'model_file': absolute_url,
                            'position_x': float(vrm.position_x),
                            'position_y': float(vrm.position_y),
                            'position_z': float(vrm.position_z),
                            'rotation_x': float(vrm.rotation_x),
                            'rotation_y': float(vrm.rotation_y),
                            'rotation_z': float(vrm.rotation_z),
                            'scale_x': float(vrm.scale_x),
                            'scale_y': float(vrm.scale_y),
                            'scale_z': float(vrm.scale_z),
                            'type': model.name  # Add type field for the frontend to identify the model
                        }
                        room_data['models'].append(model_data)
                    except Exception as e:
                        print(f"Error processing model URL: {str(e)}")

            return JsonResponse(room_data)
            
    except VirtualRoom.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Room not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



@login_required
@require_http_methods(["DELETE"])
def delete_room(request, room_id):
    try:
        room = VirtualRoom.objects.get(id=room_id, user=request.user)
        room.delete()
        return JsonResponse({
            'status': 'success',
            'message': 'Room deleted successfully'
        })
    except VirtualRoom.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Room not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)
    


@login_required
@csrf_exempt
@require_POST
def save_room(request, room_id=None):
    """
    Combined endpoint to save a virtual room from the 3D designer to the database.
    This function handles both creating new rooms and updating existing ones.
    """
    try:
        # Parse the JSON data from request body
        data = json.loads(request.body)
        
        # Extract room data
        name = data.get('name')
        if not name:
            return JsonResponse({'success': False, 'message': 'Room name is required'}, status=400)
            
        # Get or validate numeric values
        try:
            width = float(data.get('width', 0))
            length = float(data.get('length', 0))
            height = float(data.get('height', 0))
            
            if width <= 0 or length <= 0 or height <= 0:
                return JsonResponse({
                    'success': False, 
                    'message': 'Room dimensions must be positive numbers'
                }, status=400)
                
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False, 
                'message': 'Invalid room dimensions'
            }, status=400)
            
        # Get color values
        wall_color = data.get('wall_color', '#FFFFFF')
        floor_color = data.get('floor_color', '#FFFFFF')
        ceiling_color = data.get('ceiling_color', '#FFFFFF')
        
        # Get furniture data
        furniture_items = data.get('furniture', [])
        
        # Check if we're updating an existing room
        if room_id:
            try:
                room = VirtualRoom.objects.get(id=room_id, user=request.user)
                # Update existing room
                room.name = name
                room.width = width
                room.length = length
                room.height = height
                room.wall_color = wall_color
                room.floor_color = floor_color
                room.ceiling_color = ceiling_color
                room.save()
                
                # Clear existing furniture for this room to avoid duplicates
                VirtualRoomModel.objects.filter(virtual_room=room).delete()
                
                message = 'Room updated successfully'
            except VirtualRoom.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Room not found or you do not have permission to update it'
                }, status=404)
        else:
            # Check if a room with this name already exists for this user
            existing_room = VirtualRoom.objects.filter(user=request.user, name=name).first()
            
            if existing_room:
                # Update existing room with the same name
                existing_room.width = width
                existing_room.length = length
                existing_room.height = height
                existing_room.wall_color = wall_color
                existing_room.floor_color = floor_color
                existing_room.ceiling_color = ceiling_color
                existing_room.save()
                
                # Clear existing furniture for this room to avoid duplicates
                VirtualRoomModel.objects.filter(virtual_room=existing_room).delete()
                
                room = existing_room
                message = 'Room updated successfully'
            else:
                # Create new room
                room = VirtualRoom.objects.create(
                    user=request.user,
                    name=name,
                    width=width,
                    length=length,
                    height=height,
                    wall_color=wall_color,
                    floor_color=floor_color,
                    ceiling_color=ceiling_color
                )
                
                message = 'Room created successfully'
        
        # Now save all furniture items
        furniture_saved = 0
        for item in furniture_items:
            try:
                # Get or create the room model
                model_name = item.get('type', 'unknown')
                category = item.get('category') or get_category_from_name(model_name)
                
                room_model, created = RoomModel.objects.get_or_create(
                    name=model_name,
                    defaults={
                        'category': category,
                        'model_file': f'default_models/{model_name}.glb'  # Default path
                    }
                )
                
                # If model exists but category is empty, update it
                if not created and not room_model.category:
                    room_model.category = category
                    room_model.save()
                
                # Create the virtual room model entry with position data
                VirtualRoomModel.objects.create(
                    virtual_room=room,
                    room_model=room_model,
                    position_x=float(item.get('position_x', 0)),
                    position_y=float(item.get('position_y', 0)),
                    position_z=float(item.get('position_z', 0)),
                    rotation_x=float(item.get('rotation_x', 0)),
                    rotation_y=float(item.get('rotation_y', 0)),
                    rotation_z=float(item.get('rotation_z', 0)),
                    scale_x=float(item.get('scale_x', 1)),
                    scale_y=float(item.get('scale_y', 1)),
                    scale_z=float(item.get('scale_z', 1))
                )
                furniture_saved += 1
            except Exception as e:
                print(f"Error saving furniture item: {e}")
                # Continue with other items even if one fails
        
        # Return success response
        return JsonResponse({
            'success': True,
            'message': f"{message} with {furniture_saved} furniture items",
            'room_id': room.id,
            'redirect_url': '/virtual-room/'  # Redirect to room list page - update as needed
        })
        
    except Exception as e:
        # Log the error and return an error response
        print(f"Error in save_room view: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Error saving room: {str(e)}'
        }, status=500)

def get_category_from_name(name):
    """
    Infer the category of a furniture item based on its name.
    This function looks for keywords in the name and returns the appropriate category.
    """
    if not name:
        return 'furniture'
        
    name_lower = name.lower()
    
    # Define category keywords
    category_mapping = {
        'bed': 'bed',
        'sofa': 'sofa',
        'chair': 'chair',
        'table': 'table',
        'dining': 'dining',
        'desk': 'desk',
        'lamp': 'lamp',
        'light': 'lamp',
        'bookshelf': 'storage',
        'shelf': 'storage',
        'cabinet': 'storage',
        'wardrobe': 'storage',
        'dresser': 'storage',
        'nightstand': 'nightstand',
        'rug': 'rug',
        'carpet': 'rug',
        'plant': 'decor',
        'decoration': 'decor',
        'art': 'wall',
        'painting': 'wall',
        'mirror': 'wall',
        'tv': 'electronics',
        'television': 'electronics'
    }
    
    # Check for category keywords in the model name
    for keyword, category in category_mapping.items():
        if keyword in name_lower:
            return category
    
    # Default category
    return 'furniture'

